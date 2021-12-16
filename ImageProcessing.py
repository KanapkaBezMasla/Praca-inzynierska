from PIL import Image
import cv2
from PyQt5.QtWidgets import QDesktopWidget
from openpyxl import Workbook, load_workbook
import os
from WarningWindow import WarningWindow
import math



class ImageProcessing:

    @staticmethod
    def findFirstChanOnImg(chanY: int, chanN: int, pixOfMark: int, yBeg: int, yDest: int, compYPix: int):
        ymin = 155 + (compYPix - 1) * 3
        ymin = max(ymin, yBeg)
        if pixOfMark < yBeg:
            while pixOfMark < ymin:
                pixOfMark += chanY
                chanN += 1
        else:
            while pixOfMark - chanY > ymin:
                pixOfMark -= chanY
                chanN -= 1
        if pixOfMark<yBeg or pixOfMark>yDest:
            return -1, -1
        return chanN, pixOfMark


    @staticmethod
    def binarization(openFile: str, savingFile: str, threshold: int):
        img = cv2.imread(openFile)
        th, im_th = cv2.threshold(img, threshold, 255, cv2.THRESH_BINARY)
        cv2.imwrite(savingFile, im_th)

    @staticmethod
    def measurement(mmPerPix: int, chanY: int, markedChan: int, pixLine: int, xBeg: int, yBeg: int, yDest: int, compYPix: int, x_scale_val: int, x_scale_pos: int):
        #Funkcja służąca pomierzeniu uszkodzeń i zapisaniu ich do pliku

        #ustalenie x dla początku screena względem osi
        x_scale_val *= 1000
        x_scale_val += round(((xBeg - x_scale_pos)*mmPerPix)*0.67)
        chanN, pixLine = ImageProcessing.findFirstChanOnImg(chanY, markedChan, pixLine, yBeg, yDest, compYPix)
        if pixLine < 0:
            WarningWindow('Błąd odczytu osi Y! Pomiar nie wykonany!')
            return
        pixLine = pixLine - yBeg

        # otwieranie pliku
        if os.path.isfile('pomiaryUszkodzen.xlsx') and os.access('pomiaryUszkodzen.xlsx', os.W_OK):
            wb = load_workbook('pomiaryUszkodzen.xlsx')
        else:
            wb = Workbook()
        ws = wb.active
        ws.title = "dane"
        binarizated = Image.open('binarizated.png')
        width, height = binarizated.size
        yellow = True
        countingOn = False
        greenCounting = 0
        damageLen = 0
        redCounting = 0
        showedWarning = False
        showedWarning2 = False
        sheetRow = []
        emptyChan = False
        ymax = QDesktopWidget().screenGeometry().height()*2 - 90
        ymax = min(yDest, ymax)
        while pixLine + yBeg < ymax:
            # Jeżeli poprzedni wiersz kończył się w pasku, a nie w "szarej strefie"
            if countingOn:
                if yellow == True:
                    sheetRow.append('pocz: ' + str(float(x_scale_val + round(float((x - damageLen - greenCounting) * mmPerPix) * 0.67))/1000) + 'm')
                    sheetRow.append('z ' + str(round(float((damageLen + greenCounting) * mmPerPix) * 0.67)) + 'mm')
                    #ws.append()
                else:
                    sheetRow.append('pocz: ' + str(float(x_scale_val + round(float((x - damageLen - greenCounting) * mmPerPix) * 0.67))/1000) + 'm')
                    sheetRow.append('n ' + str(round(float((damageLen + greenCounting) * mmPerPix) * 0.67)) + 'mm')
                damageLen = 0
                countingOn = False
                greenCounting = 0
                redCounting = 0
                if not showedWarning:
                    WarningWindow('Możliwe ucięcie paska! Możliwie źle zapisane dane!')
                    showedWarning = True
            if sheetRow != []:
                ws.append(sheetRow)
                emptyChan = False
            else:
                # Jeżeli poprzedni kanał nie był pusty to zrobi pusty wiersz w Excelu
                if emptyChan == False:
                    ws.append(sheetRow)
                emptyChan = True
            sheetRow.clear()
            firstDamageOnChan = True
            for x in range(width):
                p = binarizated.getpixel((x, pixLine))
                if p[1] == 255:
                    # Jeśli kolor żółty
                    if p[0] == 255 and p[2] == 0:
                        if x == 0 and not showedWarning:
                            WarningWindow('Możliwe ucięcie paska! Możliwie źle zapisane dane!')
                            showedWarning = True
                        if firstDamageOnChan:
                            firstDamageOnChan = False
                            sheetRow.append('Chan' + str(chanN))
                        # Przed żółtym jest czerwony
                        if redCounting > 0:
                            if yellow:
                                damageLen += redCounting + 1
                            else:
                                yellow = True
                            redCounting = 0
                        # Przed żółtym jest też piksel żółty
                        elif countingOn == True and yellow == True and greenCounting == 0:
                            damageLen += 1
                        # Przed żółtym jest niebieski
                        elif countingOn == True and yellow == False and greenCounting == 0:
                            sheetRow.append('pocz: ' + str(float(x_scale_val + round(float((x - damageLen - greenCounting) * mmPerPix) * 0.67))/1000) + 'm')
                            sheetRow.append('n ' + str(round(float((damageLen + greenCounting) * mmPerPix) * 0.67)) + 'mm')
                            damageLen = 1
                            yellow = True
                        # Przed żółtym jest szary
                        elif countingOn == False:
                            countingOn = True
                            damageLen = 1
                            yellow = True
                        # Przed żółtym jest zielony
                        else:
                            sheetRow.append('pocz: ' + str(float(x_scale_val + round(float((x - damageLen - greenCounting) * mmPerPix) * 0.67))/1000) + 'm')
                            sheetRow.append('n ' + str(round(float((damageLen + greenCounting) * mmPerPix) * 0.67)) + 'mm')
                            damageLen = greenCounting + 1
                            greenCounting = 0
                            yellow = True
                    # Jeśli kolor niebieski
                    elif p[0] == 0 and p[2] == 255:
                        if x == 0 and not showedWarning:
                            WarningWindow('Możliwe ucięcie paska! Możliwie źle zapisane dane!')
                            showedWarning = True
                        if firstDamageOnChan:
                            firstDamageOnChan = False
                            sheetRow.append('Chan' + str(chanN))
                        if redCounting > 0:
                            if yellow == False:
                                damageLen += redCounting + 1
                            else:
                                yellow = False
                            redCounting = 0
                        # Przed niebieskim jest też piksel niebieski
                        elif countingOn == True and yellow == False and greenCounting == 0:
                            damageLen += 1
                        # Przed niebieskim jest żółty
                        elif countingOn == True and yellow == True and greenCounting == 0:
                            sheetRow.append('pocz: ' + str(float(x_scale_val + round(float((x - damageLen - greenCounting) * mmPerPix) * 0.67))/1000) + 'm')
                            sheetRow.append('z ' + str(round(float((damageLen + greenCounting) * mmPerPix) * 0.67)) + 'mm')
                            damageLen = 1
                            yellow = False
                        # Przed niebieskim jest szary
                        elif countingOn == False:
                            countingOn = True
                            damageLen = 1
                            yellow = False
                        # Przed niebieskim jest zielony
                        else:
                            yellow = False
                            sheetRow.append('pocz: ' + str(float(x_scale_val + round(float((x - damageLen - greenCounting) * mmPerPix) * 0.67))/1000) + 'm')
                            sheetRow.append('z ' + str(round(float((damageLen + greenCounting) * mmPerPix) * 0.67)) + 'mm')
                            damageLen = greenCounting + 1
                            greenCounting = 0
                    # kolor zielony
                    elif countingOn:
                        greenCounting += 1


                # kolor szary (lub jakiś błąd koloru czerwonego/niebieskiego)
                else:
                    # kolor czerwony
                    if p[0] == 255:
                        if not showedWarning2:
                            WarningWindow(
                                'Usun czerwony wskaznik z zaznaczonego pola! Możliwe błędne zapisanie danych!')
                            showedWarning2 = True
                        if countingOn:
                            redCounting += 1
                    else:
                        redCounting = 0
                    # zakończenie paska...
                    if countingOn:
                        countingOn = False
                        # ...żółtego
                        if yellow == True:
                            sheetRow.append('pocz: ' + str(float(x_scale_val + round(float((x - damageLen - greenCounting) * mmPerPix) * 0.67))/1000) + 'm')
                            sheetRow.append('z ' + str(round(float((damageLen + greenCounting) * mmPerPix) * 0.67)) + 'mm')
                            damageLen = 0
                            greenCounting = 0
                        # ...niebieskiego
                        else:
                            sheetRow.append('pocz: ' + str(float(x_scale_val + round(float((x - damageLen - greenCounting) * mmPerPix) * 0.67))/1000) + 'm')
                            sheetRow.append('n ' + str(round(float((damageLen + greenCounting) * mmPerPix) * 0.67)) + 'mm')
                            damageLen = 0
                            greenCounting = 0
            pixLine += chanY #chanY*2-1
            chanN += 1
        try:
            wb.save('pomiaryUszkodzen.xlsx')
        except Exception:
            WarningWindow('Proszę zamknąć Excela przed rozpoczęciem pomiarów! Pomiar nie zapisany!')
        return

    @staticmethod
    def binarizationMIN():
        img = Image.open('markedArea.png')
        width, height = img.size
        black_white = Image.new(mode="L", size=(width, height))

        for y in range(height):
            for x in range(width):
                p = img.getpixel((x, y))
                min_val = min(p[0], p[1], p[2])
                if min_val < 150:
                    min_val = 0
                else:
                    min_val = 255
                black_white.putpixel((x, y), min_val)
        black_white.save('chanels.png')

    #Funkcja wyliczająca rząd pierwszego kanału na obrazku.
    #Jeśli pierwszy rząd będzie poza obrazkiem, to zwróci -1
    @staticmethod
    def firstChan(yBeg: int, yDest: int, compYPix: int, chanY: int):
        chanN = math.ceil((154 - 3*compYPix - yBeg)/(-2*(chanY-1)))
        if chanN < yDest:
            chanN -= yBeg
            return chanN
        else:
            return -1

    # Funkcja wyliczająca rząd ostatniego kanału na obrazku.
    @staticmethod
    def lastChan(yBeg: int, yDest: int, compYPix: int, chanY: int):
        chanN = math.floor((154 - 3 * compYPix - yDest) / (-2 * (chanY - 1)))
        chanN -= yBeg
        return chanN
